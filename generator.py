import logging
import pandas as pd
import os
import openpyxl
import xlwt

logger = logging.getLogger(__name__)

PLACEHOLDER = "$project"

# -------------------------
# Excel reader helper (AUTO engine)
# -------------------------

def read_excel_auto(path, sheet_name=0):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        logger.info(f"Reading XLS file with xlrd: {path}")
        return pd.read_excel(path, sheet_name=sheet_name, engine="xlrd")
    else:
        logger.info(f"Reading XLSX file with openpyxl: {path}")
        return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")

# -------------------------
# Helpers
# -------------------------

def load_projects(projects_path):
    df = read_excel_auto(projects_path, sheet_name=0)

    if "PROJECT" not in df.columns:
        raise ValueError("Projects file must have a column named 'PROJECT'")

    projects = [str(p).strip() for p in df["PROJECT"].dropna().tolist() if str(p).strip()]
    if not projects:
        raise ValueError("No projects found in projects file.")

    # Deduplicate while preserving order
    seen = set()
    uniq = []
    for p in projects:
        if p not in seen:
            seen.add(p)
            uniq.append(p)
    return uniq

def row_contains_placeholder(row):
    for v in row.values:
        if isinstance(v, str) and PLACEHOLDER in v:
            return True
    return False

def replace_placeholder_in_row(row, project):
    new_row = row.copy()
    for col in new_row.index:
        v = new_row[col]
        if isinstance(v, str) and PLACEHOLDER in v:
            new_row[col] = v.replace(PLACEHOLDER, project)
    return new_row

def process_sheet(df, projects):
    df = df.copy()

    template_rows = []
    static_rows = []

    for _, row in df.iterrows():
        if row_contains_placeholder(row):
            template_rows.append(row)
        else:
            static_rows.append(row)

    generated_rows = []

    # Keep static rows
    for row in static_rows:
        generated_rows.append(row)

    # Expand template rows
    for project in projects:
        for row in template_rows:
            new_row = replace_placeholder_in_row(row, project)
            generated_rows.append(new_row)

    out_df = pd.DataFrame(generated_rows, columns=df.columns)
    return out_df

def rows_are_equal(row1, row2):
    if len(row1) != len(row2):
        return False

    for v1, v2 in zip(row1.values, row2.values):
        if pd.isna(v1) and pd.isna(v2):
            continue
        if pd.isna(v1) or pd.isna(v2):
            return False
        if v1 != v2:
            return False
    return True

def remove_duplicate_rows(existing_df, new_df):
    unique_rows = []

    for _, new_row in new_df.iterrows():
        is_duplicate = False
        for _, existing_row in existing_df.iterrows():
            if rows_are_equal(new_row, existing_row):
                is_duplicate = True
                break
        if not is_duplicate:
            unique_rows.append(new_row)

    if not unique_rows:
        return pd.DataFrame(columns=new_df.columns)

    return pd.DataFrame(unique_rows, columns=new_df.columns)

def append_to_existing_file(existing_file, new_sheets):
    logger.info("append_to_existing_file() called -> DELTA Import")

    existing_sheets = read_excel_auto(existing_file, sheet_name=None)

    merged_sheets = {}

    for sheet_name in existing_sheets.keys():
        existing_df = existing_sheets[sheet_name]

        if sheet_name in new_sheets:
            new_df = new_sheets[sheet_name]
            unique_new_df = remove_duplicate_rows(existing_df, new_df)

            if len(unique_new_df) > 0:
                logger.info(f"Sheet '{sheet_name}': appending {len(unique_new_df)} rows")
                merged_df = pd.concat([existing_df, unique_new_df], ignore_index=True)
                merged_sheets[sheet_name] = merged_df
            else:
                logger.info(f"Sheet '{sheet_name}': no new rows to append")
                merged_sheets[sheet_name] = existing_df
        else:
            merged_sheets[sheet_name] = existing_df

    # Add any new sheets
    for sheet_name in new_sheets.keys():
        if sheet_name not in existing_sheets:
            logger.info(f"Adding new sheet: {sheet_name}")
            merged_sheets[sheet_name] = new_sheets[sheet_name]

    return merged_sheets

def convert_xlsx_to_xls(xlsx_path, xls_path):

    logger.info(f"Converting XLSX to XLS: {xlsx_path} -> {xls_path}")

    wb_xlsx = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_xls = xlwt.Workbook()

    for sheet_name in wb_xlsx.sheetnames:
        ws_xlsx = wb_xlsx[sheet_name]
        ws_xls = wb_xls.add_sheet(sheet_name[:31])  # XLS sheet name limit

        for r_idx, row in enumerate(ws_xlsx.iter_rows(values_only=True)):
            for c_idx, value in enumerate(row):
                ws_xls.write(r_idx, c_idx, value)

    wb_xls.save(xls_path)
    logger.info(f"XLS file created at: {xls_path}")

# -------------------------
# Main generator function
# -------------------------

def generate_sre(template_path, projects_path, owners_path, baseline_path, output_path):
    logger.info("Starting SOD generation")

    # Load inputs
    all_sheets = read_excel_auto(template_path, sheet_name=None)
    projects = load_projects(projects_path)
    owners_df = read_excel_auto(owners_path, sheet_name=0)

    required_cols = {"PROJECT", "OWNER TYPE", "OWNER NAME", "RANK"}
    if not required_cols.issubset(set(owners_df.columns)):
        raise ValueError(f"Owners file must have columns: {required_cols}")

    logger.info(f"Loaded template with {len(all_sheets)} sheets")
    logger.info(f"Loaded {len(projects)} projects")
    logger.info(f"Loaded owners file with {len(owners_df)} rows")

    # Generate all sheets normally
    new_sheets = {}
    for sheet_name, df in all_sheets.items():
        logger.info(f"Processing sheet: {sheet_name}")
        out_df = process_sheet(df, projects)
        new_sheets[sheet_name] = out_df

    # -------------------------
    # Build Risks-Owners from Risks × Owners
    # -------------------------
    risks_sheet_name = None
    for name in new_sheets.keys():
        key = name.strip().lower().replace(" ", "").replace("-", "")
        if key == "risks":
            risks_sheet_name = name
            break

    if risks_sheet_name:
        logger.info(f"Building Risks-Owners sheet from: {risks_sheet_name}")
        risks_df = new_sheets[risks_sheet_name]
        if "RISK NAME" not in risks_df.columns:
            raise ValueError("Risks sheet must have a 'RISK NAME' column")

        rows = []
        for _, risk_row in risks_df.iterrows():
            risk_name = risk_row["RISK NAME"]
            if not isinstance(risk_name, str):
                continue

            proj = None
            if "(" in risk_name and ")" in risk_name:
                proj = risk_name.split("(")[-1].split(")")[0].strip()

            if not proj:
                continue

            owners_for_project = owners_df[
                owners_df["PROJECT"].astype(str).str.strip() == proj
            ]

            for _, owner_row in owners_for_project.iterrows():
                rows.append({
                    "RISK NAME": risk_name,
                    "OWNER TYPE": owner_row["OWNER TYPE"],
                    "OWNER NAME": owner_row["OWNER NAME"],
                    "RANK": owner_row["RANK"],
                })

        risk_owners_df = pd.DataFrame(
            rows, columns=["RISK NAME", "OWNER TYPE", "OWNER NAME", "RANK"]
        )
        logger.info(f"Generated Risks-Owners sheet with {len(risk_owners_df)} rows")
        new_sheets["Risks-Owners"] = risk_owners_df

    # -------------------------
    # Baseline (delta) or full
    # -------------------------
    use_delta = False

    if baseline_path and os.path.exists(baseline_path):
        logger.info(f"Output file provided: {baseline_path}")
        try:
            read_excel_auto(baseline_path, sheet_name=None)
            use_delta = True
        except Exception as e:
            logger.warning("Failed to read output file, switching to FULL Import", exc_info=e)
            use_delta = False
    else:
        logger.info("No Output file provided")

    if use_delta:
        logger.info("Mode: DELTA Import")
        final_sheets = append_to_existing_file(baseline_path, new_sheets)
    else:
        logger.info("Mode: FULL Import (no valid output file found)")
        final_sheets = new_sheets

    # Write output XLSX
    logger.info("Writing output XLSX file")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, out_df in final_sheets.items():
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)

    logger.info(f"XLSX file written at: {output_path}")

    # Convert to XLS
    xls_path = output_path.replace(".xlsx", ".xls")
    convert_xlsx_to_xls(output_path, xls_path)
